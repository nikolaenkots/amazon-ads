from bq_client import get_client
import csv
import os
import re
import threading
from datetime import datetime, timezone
from flask import Blueprint, request, jsonify, send_from_directory
from google.cloud import bigquery
from google.cloud.bigquery import LoadJobConfig

catalog_bp = Blueprint('catalog', __name__)

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
PROJECT_ID    = "amazon-ads-api-494412"
DATASET       = "amazon_ads"
TABLE         = "catalog"
STAGING       = "catalog_staging"
CHUNK_SIZE    = 1000

# Импортируем shared progress_store из app
def _get_progress_store():
    import app
    return app.progress_store

def emit(job_id, event, data):
    ps = _get_progress_store()
    if job_id not in ps:
        ps[job_id] = []
    ps[job_id].append({"event": event, "data": data})

def load_chunk(client, table_ref, chunk):
    job = client.load_table_from_json(
        chunk, table_ref,
        job_config=LoadJobConfig(write_disposition="WRITE_APPEND")
    )
    job.result()
    return job.errors or []

def parse_ts(ts):
    try:
        if ts and float(ts) > 0:
            return datetime.fromtimestamp(float(ts), tz=timezone.utc).isoformat()
    except:
        pass
    return None

def parse_bullets(json_part):
    idx_s = json_part.find('bullets":[')
    idx_e = json_part.find(',description:', idx_s)
    if idx_s < 0 or idx_e < 0:
        return None, None
    bc = json_part[idx_s + len('bullets":['):idx_e]
    if not bc.strip() or bc.strip() == ']':
        return None, None
    bullets = []
    pos = 0
    while pos < len(bc):
        c = bc[pos]
        if c == '"':
            end = pos + 1
            while end < len(bc):
                if bc[end] == '"':
                    n = 0; k = end - 1
                    while k >= 0 and bc[k] == '\\': n += 1; k -= 1
                    if n % 2 == 0: break
                end += 1
            item = bc[pos+1:end].replace('\\"', '"')
            bullets.append(item.strip())
            pos = end + 1
            if pos < len(bc) and bc[pos] == ',': pos += 1
        elif c in (',', ']'):
            pos += 1
        else:
            nq = bc.find(',"', pos)
            item = (bc[pos:nq] if nq > 0 else bc[pos:]).strip().strip(']')
            if item: bullets.append(item)
            break
    return (bullets[0] if len(bullets) > 0 else None,
            bullets[1] if len(bullets) > 1 else None)

def parse_pm(row):
    jp = ','.join(row[17:])
    b1, b2 = parse_bullets(jp)
    img = None
    m = re.search(r'(https://m\.media-amazon\.com/images/[^\s",}\\]+)', jp)
    if m: img = m.group(1)
    return b1, b2, img

def build_edit_url(design_id, listing_id, mp):
    try:
        suffix = listing_id.replace(design_id + '_', '')
        pt = '_'.join(suffix.split('_')[:-1])
        return (f"https://merch.amazon.com/designs/{design_id}/edit"
                f"?productor-product-type={pt}&productor-marketplace={mp}")
    except:
        return None

# ── Productor xlsx export ─────────────────────────────────
def _clean(v):
    """Значение ячейки xlsx → str; None и литерал 'None' → пустая строка."""
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s == "None" else s

# Отображаемое имя ProductType → внутренний токен Productor.
# Токены обязаны совпадать со старым каталогом (listing_id, product_type),
# иначе MERGE задвоит строки, а копирование кампаний перестанет матчиться.
# Первичный источник токена — колонка fileName ({Title}-{TOKEN}-{ASIN}-productor.png),
# эта таблица — fallback, если fileName не распарсился.
PT_MAP = {
    "Standard T-Shirt":          "STANDARD_TSHIRT",
    "Premium T-Shirt":           "PREMIUM_TSHIRT",
    "V-neck T-Shirt":            "VNECK",
    "Long Sleeve T-Shirt":       "STANDARD_LONG_SLEEVE",
    "Sweatshirt":                "STANDARD_SWEATSHIRT",
    "Pullover Hoodie":           "STANDARD_PULLOVER_HOODIE",
    "Zip Hoodie":                "ZIP_HOODIE",
    "Tank Top":                  "TANK_TOP",
    "Raglan":                    "RAGLAN",
    "Mug":                       "MUG",
    "Tumbler":                   "TUMBLER",
    "Water Bottle":              "WATER_BOTTLE",
    "OVERSIZED_TSHIRT":          "OVERSIZED_TSHIRT",
    "PERFORMANCE_TSHIRT":        "PERFORMANCE_TSHIRT",
    "COMFORT_COLORS_SWEATSHIRT": "COMFORT_COLORS_SWEATSHIRT",
}

def _pt_token(rec):
    """Токен типа продукта: из fileName → из PT_MAP → эвристика."""
    fn   = _clean(rec.get("fileName"))
    asin = _clean(rec.get("Asin"))
    if fn and asin:
        m = re.search(r'-([A-Z0-9_]+)-' + re.escape(asin) + r'-productor\.png$', fn)
        if m:
            return m.group(1)
    pt = _clean(rec.get("ProductType"))
    if pt in PT_MAP:
        return PT_MAP[pt]
    return re.sub(r'[^A-Z0-9]+', '_', pt.upper()).strip('_')

def parse_productor_dt(v):
    """Created из Productor: datetime-ячейка или строка '04/30/2024 6:36 AM'."""
    if isinstance(v, datetime):
        return v.replace(tzinfo=timezone.utc).isoformat()
    s = _clean(v)
    if not s:
        return None
    for fmt in ("%m/%d/%Y %I:%M %p", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc).isoformat()
        except ValueError:
            pass
    return None

def process_productor_row(rec):
    """rec — dict {заголовок: значение} строки Productor-экспорта."""
    asin = _clean(rec.get("Asin"))
    mp   = _clean(rec.get("Marketplace"))
    did  = _clean(rec.get("DesignId"))
    pt   = _pt_token(rec)
    if not asin or not mp or not did or not pt:
        return None
    ad_asin = _clean(rec.get("Default Color Asin (Variation for ADs)"))
    if not ad_asin and pt == "STANDARD_TSHIRT":
        ad_asin = asin
    price = _clean(rec.get("Price"))
    try:
        price = float(price) if price else None
    except ValueError:
        price = None
    return {
        "listing_id":        f"{did}_{pt}_{mp}",
        "asin":              asin,
        "marketplace":       mp,
        "design_id":         did,
        "brand":             _clean(rec.get("Brand")) or None,
        "title":             _clean(rec.get("Title")) or None,
        "product_type":      pt,
        "price":             price,
        "status":            _clean(rec.get("Status")) or None,
        "bullet_point_1":    _clean(rec.get("Bullet Points 1")) or None,
        "bullet_point_2":    _clean(rec.get("Bullet Points 2")) or None,
        "ad_asin":           ad_asin or None,
        "image_url":         _clean(rec.get("Mockup Url")) or None,
        "live_url":          _clean(rec.get("Live Url")) or f"https://www.amazon.com/dp/{asin}",
        "edit_url":          _clean(rec.get("Edit Url")) or build_edit_url(did, f"{did}_{pt}_{mp}", mp),
        "created_at_amazon": parse_productor_dt(rec.get("Created")),
        "imported_at":       datetime.now(tz=timezone.utc).isoformat(),
    }

def iter_productor_rows(filepath):
    """Читает Productor xlsx, отдаёт dict'ы {заголовок: значение}."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("Для импорта .xlsx нужен пакет openpyxl (pip install openpyxl)")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [(_clean(h)) for h in next(rows, [])]
    if "Asin" not in header or "DesignId" not in header:
        raise RuntimeError("Не похоже на экспорт Productor: нет колонок Asin / DesignId")
    for r in rows:
        yield dict(zip(header, r))
    wb.close()


def process_catalog_row(row):
    if len(row) < 17: return None
    asin = row[0].strip()
    mp   = row[10].strip()
    lid  = row[8].strip()
    if not asin or not mp or not lid: return None
    did = row[5].strip()
    b1, b2, img = parse_pm(row)
    return {
        "listing_id":        lid,
        "asin":              asin,
        "marketplace":       mp,
        "design_id":         did,
        "brand":             row[1].strip() or None,
        "title":             row[12].strip() or None,
        "product_type":      row[13].strip() or None,
        "price":             float(row[7]) if row[7].strip() else None,
        "status":            row[15].strip() or None,
        "bullet_point_1":    b1,
        "bullet_point_2":    b2,
        "ad_asin":           asin if row[13].strip() == "STANDARD_T_SHIRT" else None,
        "image_url":         img,
        "live_url":          f"https://www.amazon.com/dp/{asin}",
        "edit_url":          build_edit_url(did, lid, mp),
        "created_at_amazon": parse_ts(row[2]),
        "imported_at":       datetime.now(tz=timezone.utc).isoformat(),
    }

def recreate_staging(client, staging_ref):
    client.query(f"DROP TABLE IF EXISTS `{staging_ref}`").result()
    client.query(f"""
        CREATE TABLE `{staging_ref}` (
            listing_id STRING NOT NULL,
            asin STRING NOT NULL,
            marketplace STRING NOT NULL,
            design_id STRING,
            brand STRING,
            title STRING,
            product_type STRING,
            price FLOAT64,
            status STRING,
            bullet_point_1 STRING,
            bullet_point_2 STRING,
            ad_asin STRING,
            image_url STRING,
            live_url STRING,
            edit_url STRING,
            created_at_amazon TIMESTAMP,
            imported_at TIMESTAMP
        ) CLUSTER BY marketplace, product_type, status
    """).result()

def run_catalog_import(filepath, job_id):
    errors_log = []
    try:
        client      = get_client()
        staging_ref = f"{PROJECT_ID}.{DATASET}.{STAGING}"
        table_ref   = f"{PROJECT_ID}.{DATASET}.{TABLE}"

        emit(job_id, "step", {"step": 1, "msg": "Пересоздаём staging таблицу..."})
        recreate_staging(client, staging_ref)

        is_xlsx = filepath.lower().endswith('.xlsx')
        emit(job_id, "step", {"step": 2, "msg": f"Загружаем данные в staging ({'Productor xlsx' if is_xlsx else 'CSV'})..."})
        total = inserted = err_count = 0
        chunk = []

        def _records():
            """Генератор готовых записей staging-схемы из файла любого формата."""
            if is_xlsx:
                for rec in iter_productor_rows(filepath):
                    yield process_productor_row(rec)
            else:
                with open(filepath, 'r', encoding='utf-8', newline='') as f:
                    reader = csv.reader(f)
                    next(reader, None)
                    for row in reader:
                        if not row or len(row) < 17:
                            continue
                        yield process_catalog_row(row)

        for rec in _records():
                if rec:
                    chunk.append(rec)
                    total += 1
                else:
                    err_count += 1
                if len(chunk) >= CHUNK_SIZE:
                    errs = load_chunk(client, staging_ref, chunk)
                    if errs:
                        err_count += len(errs)
                        for e in errs[:3]: errors_log.append(str(e))
                    else:
                        inserted += len(chunk)
                    chunk = []
                    emit(job_id, "progress", {
                        "total": total, "inserted": inserted,
                        "errors": err_count,
                        "pct": min(75, int(inserted / max(total, 1) * 75))
                    })

        if chunk:
            errs = load_chunk(client, staging_ref, chunk)
            if errs:
                err_count += len(errs)
                for e in errs[:3]: errors_log.append(str(e))
            else:
                inserted += len(chunk)

        emit(job_id, "progress", {"total": total, "inserted": inserted, "errors": err_count, "pct": 80})
        emit(job_id, "step", {"step": 3, "msg": "Запускаем MERGE в основную таблицу..."})

        merge_sql = f"""
        MERGE `{table_ref}` AS target
        USING `{staging_ref}` AS source
          ON target.listing_id = source.listing_id
        WHEN MATCHED THEN UPDATE SET
          target.asin=source.asin, target.marketplace=source.marketplace,
          target.design_id=source.design_id, target.brand=source.brand,
          target.title=source.title, target.product_type=source.product_type,
          target.price=source.price, target.status=source.status,
          target.bullet_point_1=source.bullet_point_1,
          target.bullet_point_2=source.bullet_point_2,
          target.ad_asin=source.ad_asin, target.image_url=source.image_url,
          target.live_url=source.live_url, target.edit_url=source.edit_url,
          target.imported_at=source.imported_at
        WHEN NOT MATCHED THEN INSERT (
          listing_id, asin, marketplace, design_id, brand, title, product_type,
          price, status, bullet_point_1, bullet_point_2, ad_asin,
          image_url, live_url, edit_url, created_at_amazon, imported_at
        ) VALUES (
          source.listing_id, source.asin, source.marketplace, source.design_id,
          source.brand, source.title, source.product_type, source.price,
          source.status, source.bullet_point_1, source.bullet_point_2,
          source.ad_asin, source.image_url, source.live_url, source.edit_url,
          source.created_at_amazon, source.imported_at
        )
        """
        client.query(merge_sql).result()

        emit(job_id, "progress", {"total": total, "inserted": inserted, "errors": err_count, "pct": 95})
        emit(job_id, "step", {"step": 4, "msg": "Очищаем staging..."})
        recreate_staging(client, staging_ref)

        cnt = list(client.query(f"SELECT COUNT(*) as cnt FROM `{table_ref}`").result())[0].cnt
        emit(job_id, "done", {
            "total": total, "inserted": inserted, "errors": err_count,
            "total_in_table": cnt, "errors_log": errors_log[:20], "pct": 100
        })
        try: os.remove(filepath)
        except: pass

    except Exception as e:
        emit(job_id, "error", {"msg": str(e)})


@catalog_bp.route('/catalog')
def catalog_page():
    return send_from_directory(BASE_DIR, 'catalog.html')

@catalog_bp.route('/upload', methods=['POST'])
def upload():
    f = request.files.get('file')
    if not f:
        return jsonify({"error": "Файл не найден"}), 400
    job_id   = datetime.now().strftime('%Y%m%d%H%M%S%f')
    ext      = '.xlsx' if (f.filename or '').lower().endswith('.xlsx') else '.csv'
    filepath = os.path.join(UPLOAD_FOLDER, f'catalog_{job_id}{ext}')
    f.save(filepath)
    ps = _get_progress_store()
    ps[job_id] = []
    threading.Thread(target=run_catalog_import, args=(filepath, job_id), daemon=True).start()
    return jsonify({"job_id": job_id})

@catalog_bp.route('/progress/<job_id>')
def progress(job_id):
    ps = _get_progress_store()
    if job_id not in ps:
        return jsonify([])
    msgs = list(ps[job_id])
    ps[job_id] = []
    return jsonify(msgs)